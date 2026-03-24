"""
Pro Forma spreadsheet generator.
Creates a multi-tab xlsx with property inputs, summary metrics,
5-year projections, comps placeholder, CapEx budget, and tax projections.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers


# ── Style constants ─────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
_SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_LABEL_FONT = Font(name="Calibri", bold=True, size=10)
_BODY_FONT = Font(name="Calibri", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_DOLLAR_FMT = '#,##0'
_PCT_FMT = '0.0%'


def _style_header_row(ws, row, max_col):
    """Apply dark blue header styling to an entire row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _style_cell(ws, row, col, value, fmt=None, bold=False):
    """Write a value and apply standard formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _LABEL_FONT if bold else _BODY_FONT
    cell.border = _THIN_BORDER
    if fmt == "dollar":
        cell.number_format = _DOLLAR_FMT
    elif fmt == "pct":
        cell.number_format = _PCT_FMT
    return cell


def _safe(d, key, default=None):
    """Safely get a value from a dict."""
    v = d.get(key, default)
    return v if v is not None else default


# ─────────────────────────────────────────────────────────────────
# TAB BUILDERS
# ─────────────────────────────────────────────────────────────────

def _build_inputs_tab(ws, d):
    """Tab 1 -- Property Info, Unit Mix, Acquisition Assumptions."""
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # ── Property Information ────────────────────────────────────
    r = 1
    ws.cell(row=r, column=1, value="PROPERTY INFORMATION")
    _style_header_row(ws, r, 2)
    r += 1
    info_rows = [
        ("Property Name", _safe(d, "property_name", "")),
        ("Address", _safe(d, "address", "")),
        ("City", _safe(d, "city", "")),
        ("State", _safe(d, "state", "TX")),
        ("Zip", _safe(d, "zip", "")),
        ("Units", _safe(d, "units", 0)),
        ("Total SF", _safe(d, "sqft", 0)),
        ("Year Built", _safe(d, "year_built", "")),
        ("Property Type", "Multifamily"),
    ]
    for label, val in info_rows:
        _style_cell(ws, r, 1, label, bold=True)
        _style_cell(ws, r, 2, val)
        r += 1

    # ── Unit Mix ────────────────────────────────────────────────
    r += 1
    ws.cell(row=r, column=1, value="UNIT MIX")
    _style_header_row(ws, r, 5)
    r += 1
    for col, hdr in enumerate(["Type", "Count", "Avg SF", "Current Rent", "Market Rent"], 1):
        _style_cell(ws, r, col, hdr, bold=True)
    r += 1

    unit_mix = _safe(d, "unit_mix", [])
    if unit_mix:
        for unit in unit_mix:
            _style_cell(ws, r, 1, unit.get("type", ""))
            _style_cell(ws, r, 2, unit.get("count", 0))
            _style_cell(ws, r, 3, unit.get("avg_sf", 0))
            _style_cell(ws, r, 4, unit.get("current_rent", 0), "dollar")
            _style_cell(ws, r, 5, unit.get("market_rent", 0), "dollar")
            r += 1
    else:
        _style_cell(ws, r, 1, "All Units")
        _style_cell(ws, r, 2, _safe(d, "units", 0))
        _style_cell(ws, r, 3, "")
        avg_rent = 0
        units = _safe(d, "units", 1) or 1
        gpr = _safe(d, "gross_potential_rent_annual", 0)
        if gpr:
            avg_rent = gpr / 12 / units
        _style_cell(ws, r, 4, round(avg_rent), "dollar")
        _style_cell(ws, r, 5, "", "dollar")
        r += 1

    # ── Acquisition Assumptions ─────────────────────────────────
    r += 1
    ws.cell(row=r, column=1, value="ACQUISITION ASSUMPTIONS")
    _style_header_row(ws, r, 2)
    r += 1
    acq_rows = [
        ("Purchase Price", _safe(d, "price", 0), "dollar"),
        ("Price Per Unit", _safe(d, "price_per_unit", 0), "dollar"),
        ("Down Payment %", _safe(d, "down_payment_pct", 0.20), "pct"),
        ("Down Payment $", _safe(d, "down_payment", 0), "dollar"),
        ("Loan Amount", _safe(d, "loan_amount", 0), "dollar"),
        ("Interest Rate", _safe(d, "interest_rate", 0.075), "pct"),
        ("Amortization (Yrs)", _safe(d, "amortization_years", 30), None),
        ("Closing Costs (est.)", round(_safe(d, "price", 0) * 0.02), "dollar"),
    ]
    for label, val, fmt in acq_rows:
        _style_cell(ws, r, 1, label, bold=True)
        _style_cell(ws, r, 2, val, fmt)
        r += 1


def _build_summary_tab(ws, d):
    """Tab 2 -- Key Metrics dashboard."""
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    r = 1
    ws.cell(row=r, column=1, value="KEY METRICS SUMMARY")
    _style_header_row(ws, r, 2)
    r += 1

    metrics = [
        ("Purchase Price", _safe(d, "price", 0), "dollar"),
        ("Units", _safe(d, "units", 0), None),
        ("Price Per Unit", _safe(d, "price_per_unit", 0), "dollar"),
        ("", "", None),
        ("Net Operating Income (T-12)", _safe(d, "noi", 0), "dollar"),
        ("Cap Rate (As-Is)", _safe(d, "cap_rate", 0), "pct"),
        ("Cash-on-Cash (As-Is)", _safe(d, "cash_on_cash", 0), "pct"),
        ("DSCR", _safe(d, "dscr", 0), None),
        ("GRM", _safe(d, "grm", 0), None),
        ("", "", None),
        ("Value-Add NOI", _safe(d, "va_noi", 0), "dollar"),
        ("Value-Add Cap Rate", _safe(d, "va_cap_rate", 0), "pct"),
        ("Value-Add CoC", _safe(d, "va_cash_on_cash", 0), "pct"),
        ("", "", None),
        ("5-Year IRR", _safe(d, "irr_5yr", 0), "pct"),
        ("Equity Multiple", _safe(d, "equity_multiple_5yr", 0), None),
        ("Exit Value (5yr)", _safe(d, "exit_value", 0), "dollar"),
        ("Total Profit (5yr)", _safe(d, "total_profit_5yr", 0), "dollar"),
        ("", "", None),
        ("Verdict", _safe(d, "verdict", ""), None),
    ]
    for label, val, fmt in metrics:
        _style_cell(ws, r, 1, label, bold=True)
        _style_cell(ws, r, 2, val, fmt)
        r += 1


def _build_pro_forma_tab(ws, d):
    """Tab 3 -- 5-year projection (T-12 through Year 5)."""
    ws.title = "Pro Forma"
    ws.column_dimensions["A"].width = 32
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 16

    rent_growth = 0.03
    expense_growth = 0.03

    units = _safe(d, "units", 1) or 1
    noi = _safe(d, "noi", 0)
    egi = _safe(d, "effective_gross_income", 0)
    gpr = _safe(d, "gross_potential_rent_annual", 0)
    opex = _safe(d, "total_operating_expenses", 0)
    vacancy_rate = _safe(d, "vacancy_rate", 0.07)
    debt_service = _safe(d, "annual_loan_payment", 0)

    # Header row
    r = 1
    headers = ["", "T-12", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for c, hdr in enumerate(headers, 1):
        _style_cell(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, 7)

    # Income section
    r += 1
    ws.cell(row=r, column=1, value="INCOME")
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = _SUBHEADER_FILL
        ws.cell(row=r, column=c).font = _SUBHEADER_FONT
    r += 1

    income_rows = [
        ("Gross Potential Rent", gpr),
        ("Less: Vacancy", -(gpr * vacancy_rate) if gpr else 0),
        ("Effective Gross Income", egi),
    ]
    for label, t12_val in income_rows:
        _style_cell(ws, r, 1, label, bold=True)
        _style_cell(ws, r, 2, round(t12_val), "dollar")
        for yr in range(1, 6):
            projected = t12_val * ((1 + rent_growth) ** yr)
            _style_cell(ws, r, 2 + yr, round(projected), "dollar")
        r += 1

    # Expenses section
    r += 1
    ws.cell(row=r, column=1, value="OPERATING EXPENSES")
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = _SUBHEADER_FILL
        ws.cell(row=r, column=c).font = _SUBHEADER_FONT
    r += 1

    # Break down expenses
    price = _safe(d, "price", 0)
    taxes = price * 0.021
    insurance = units * 600
    management = egi * 0.08
    maintenance = units * 1200
    capex_reserve = units * 600
    admin = units * 300

    expense_items = [
        ("Property Taxes", taxes),
        ("Insurance", insurance),
        ("Management (8%)", management),
        ("Maintenance", maintenance),
        ("CapEx Reserve", capex_reserve),
        ("Admin / G&A", admin),
        ("Total Operating Expenses", opex),
    ]
    for label, t12_val in expense_items:
        is_total = label.startswith("Total")
        _style_cell(ws, r, 1, label, bold=is_total)
        _style_cell(ws, r, 2, round(t12_val), "dollar")
        for yr in range(1, 6):
            projected = t12_val * ((1 + expense_growth) ** yr)
            _style_cell(ws, r, 2 + yr, round(projected), "dollar")
        r += 1

    # NOI & Cash Flow
    r += 1
    ws.cell(row=r, column=1, value="NET OPERATING INCOME")
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = _SUBHEADER_FILL
        ws.cell(row=r, column=c).font = _SUBHEADER_FONT
    r += 1

    _style_cell(ws, r, 1, "NOI", bold=True)
    _style_cell(ws, r, 2, round(noi), "dollar")
    for yr in range(1, 6):
        proj_egi = egi * ((1 + rent_growth) ** yr)
        proj_opex = opex * ((1 + expense_growth) ** yr)
        _style_cell(ws, r, 2 + yr, round(proj_egi - proj_opex), "dollar")
    r += 1

    _style_cell(ws, r, 1, "Less: Debt Service", bold=True)
    for c in range(2, 8):
        _style_cell(ws, r, c, round(-debt_service), "dollar")
    r += 1

    _style_cell(ws, r, 1, "Cash Flow Before Tax", bold=True)
    _style_cell(ws, r, 2, round(noi - debt_service), "dollar")
    for yr in range(1, 6):
        proj_egi = egi * ((1 + rent_growth) ** yr)
        proj_opex = opex * ((1 + expense_growth) ** yr)
        proj_noi = proj_egi - proj_opex
        _style_cell(ws, r, 2 + yr, round(proj_noi - debt_service), "dollar")
    r += 1


def _build_comps_tab(ws, d):
    """Tab 4 -- Placeholder comps table."""
    ws.title = "Comps"
    ws.column_dimensions["A"].width = 6
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 18

    r = 1
    headers = ["#", "Property Name", "Address", "Units", "Year Built",
               "Sale Price", "Price/Unit", "Cap Rate", "Sale Date"]
    for c, hdr in enumerate(headers, 1):
        _style_cell(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, len(headers))

    comps = _safe(d, "comps", [])
    for i in range(5):
        r += 1
        if i < len(comps):
            comp = comps[i]
            _style_cell(ws, r, 1, i + 1)
            _style_cell(ws, r, 2, comp.get("name", ""))
            _style_cell(ws, r, 3, comp.get("address", ""))
            _style_cell(ws, r, 4, comp.get("units", ""))
            _style_cell(ws, r, 5, comp.get("year_built", ""))
            _style_cell(ws, r, 6, comp.get("sale_price", ""), "dollar")
            _style_cell(ws, r, 7, comp.get("price_per_unit", ""), "dollar")
            _style_cell(ws, r, 8, comp.get("cap_rate", ""), "pct")
            _style_cell(ws, r, 9, comp.get("sale_date", ""))
        else:
            _style_cell(ws, r, 1, i + 1)
            for c in range(2, 10):
                _style_cell(ws, r, c, "")


def _build_capex_tab(ws, d):
    """Tab 5 -- Placeholder CapEx budget."""
    ws.title = "CapEx"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 30

    r = 1
    headers = ["Item", "Cost/Unit", "Total Cost", "Notes"]
    for c, hdr in enumerate(headers, 1):
        _style_cell(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, len(headers))

    units = _safe(d, "units", 1) or 1
    reno_per_unit = _safe(d, "reno_cost_per_unit", 8000)

    items = [
        ("Interior Renovations", reno_per_unit, "Flooring, paint, fixtures, appliances"),
        ("Exterior / Curb Appeal", 1500, "Paint, landscaping, signage"),
        ("HVAC Upgrades", 1000, "As needed per unit"),
        ("Plumbing", 500, "Common area + unit updates"),
        ("Electrical", 400, "Panel upgrades, lighting"),
        ("Roof Reserve", 800, "Pro-rated replacement"),
        ("Parking / Paving", 300, "Seal-coat, striping"),
        ("Common Areas", 500, "Clubhouse, laundry, pool"),
    ]
    for item_name, cost_per_unit, notes in items:
        r += 1
        _style_cell(ws, r, 1, item_name, bold=True)
        _style_cell(ws, r, 2, cost_per_unit, "dollar")
        _style_cell(ws, r, 3, cost_per_unit * units, "dollar")
        _style_cell(ws, r, 4, notes)

    r += 2
    _style_cell(ws, r, 1, "TOTAL CAPEX BUDGET", bold=True)
    total = sum(c for _, c, _ in items) * units
    _style_cell(ws, r, 3, total, "dollar")


def _build_tax_tab(ws, d):
    """Tab 6 -- Placeholder tax projections."""
    ws.title = "Tax"
    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 16

    r = 1
    headers = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for c, hdr in enumerate(headers, 1):
        _style_cell(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, len(headers))

    price = _safe(d, "price", 0)
    # Assume 80% building / 20% land for depreciation
    building_value = price * 0.80
    annual_depreciation = building_value / 27.5  # residential 27.5 year

    tax_rows = [
        "NOI (from Pro Forma)",
        "Less: Mortgage Interest",
        "Less: Depreciation",
        "Taxable Income",
        "Estimated Tax (25%)",
        "After-Tax Cash Flow",
    ]
    r += 1
    for label in tax_rows:
        _style_cell(ws, r, 1, label, bold=True)
        for c in range(2, 7):
            _style_cell(ws, r, c, "TBD")
        r += 1

    r += 1
    _style_cell(ws, r, 1, "Annual Depreciation", bold=True)
    _style_cell(ws, r, 2, round(annual_depreciation), "dollar")
    r += 1
    _style_cell(ws, r, 1, "Building Value (80%)", bold=True)
    _style_cell(ws, r, 2, round(building_value), "dollar")


# ─────────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────────

def generate(deal_data: dict, filepath: str) -> str:
    """
    Generate a Pro Forma xlsx workbook and save it to *filepath*.

    Parameters
    ----------
    deal_data : dict
        Property and financial data for the deal.
    filepath : str
        Destination path for the .xlsx file.

    Returns
    -------
    str  The filepath that was written.
    """
    wb = Workbook()

    # Tab 1 - Inputs (uses the default sheet)
    _build_inputs_tab(wb.active, deal_data)

    # Tab 2 - Summary
    _build_summary_tab(wb.create_sheet(), deal_data)

    # Tab 3 - Pro Forma
    _build_pro_forma_tab(wb.create_sheet(), deal_data)

    # Tab 4 - Comps
    _build_comps_tab(wb.create_sheet(), deal_data)

    # Tab 5 - CapEx
    _build_capex_tab(wb.create_sheet(), deal_data)

    # Tab 6 - Tax
    _build_tax_tab(wb.create_sheet(), deal_data)

    wb.save(filepath)
    return filepath
