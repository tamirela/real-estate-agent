"""
Sensitivity Test spreadsheet generator.
Creates a multi-tab xlsx with purchase price sensitivity analysis,
monthly projection pro forma, year-by-year MOIC projections,
comps, and CapEx placeholder.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ── Style constants ─────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
_SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_GREEN_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
_RED_FILL = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
_LABEL_FONT = Font(name="Calibri", bold=True, size=10)
_BODY_FONT = Font(name="Calibri", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_DOLLAR_FMT = '#,##0'
_PCT_FMT = '0.0%'
_MOIC_FMT = '0.00"x"'


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _sc(ws, row, col, value, fmt=None, bold=False):
    """Style cell helper."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _LABEL_FONT if bold else _BODY_FONT
    cell.border = _THIN_BORDER
    if fmt == "dollar":
        cell.number_format = _DOLLAR_FMT
    elif fmt == "pct":
        cell.number_format = _PCT_FMT
    elif fmt == "moic":
        cell.number_format = _MOIC_FMT
    return cell


def _safe(d, key, default=None):
    v = d.get(key, default)
    return v if v is not None else default


# ─────────────────────────────────────────────────────────────────
# TAB 1 -- Summary: Purchase Price Sensitivity
# ─────────────────────────────────────────────────────────────────

def _build_summary_tab(ws, d):
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 18

    price = _safe(d, "price", 0)
    noi = _safe(d, "noi", 0)
    units = _safe(d, "units", 1) or 1
    down_pct = _safe(d, "down_payment_pct", 0.20)
    interest_rate = _safe(d, "interest_rate", 0.075)
    moic_target = 2.2

    # Test 5 price points: -10%, -5%, asking, +5%, +10%
    price_points = [
        ("Asking -10%", round(price * 0.90)),
        ("Asking -5%", round(price * 0.95)),
        ("Asking Price", round(price)),
        ("Asking +5%", round(price * 1.05)),
        ("Asking +10%", round(price * 1.10)),
    ]

    r = 1
    ws.cell(row=r, column=1, value=f"PURCHASE PRICE SENSITIVITY  |  MOIC Target: {moic_target}x")
    _style_header_row(ws, r, 8)

    r += 1
    header_labels = ["Scenario", "Price", "Price/Unit", "Down Payment",
                     "Cap Rate", "Year-1 CoC", "5yr MOIC (est.)", "Meets Target?"]
    for c, hdr in enumerate(header_labels, 1):
        _sc(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, 8)

    for label, test_price in price_points:
        r += 1
        cap = noi / test_price if test_price > 0 else 0
        ppu = test_price / units
        dp = test_price * down_pct
        loan = test_price - dp
        monthly_rate = interest_rate / 12
        n_payments = 30 * 12
        if monthly_rate > 0:
            monthly_pmt = loan * (monthly_rate * (1 + monthly_rate) ** n_payments) / (
                (1 + monthly_rate) ** n_payments - 1
            )
        else:
            monthly_pmt = loan / n_payments
        annual_ds = monthly_pmt * 12
        yr1_cf = noi - annual_ds
        yr1_coc = yr1_cf / dp if dp > 0 else 0

        # Rough 5yr MOIC estimate (NOI grows 3%, exit at 6.5% cap)
        exit_cap = _safe(d, "exit_cap_rate", 0.065)
        total_cf = 0
        for yr in range(1, 6):
            total_cf += (noi * (1.03 ** yr)) - annual_ds
        exit_noi = noi * (1.03 ** 5)
        exit_val = exit_noi / exit_cap if exit_cap > 0 else 0
        remaining = _remaining_balance(loan, interest_rate, 30, 5)
        net_exit = exit_val * 0.96 - remaining
        moic = (total_cf + net_exit) / dp if dp > 0 else 0

        _sc(ws, r, 1, label, bold=True)
        _sc(ws, r, 2, test_price, "dollar")
        _sc(ws, r, 3, round(ppu), "dollar")
        _sc(ws, r, 4, round(dp), "dollar")
        _sc(ws, r, 5, cap, "pct")
        _sc(ws, r, 6, yr1_coc, "pct")
        _sc(ws, r, 7, round(moic, 2), "moic")
        meets_cell = _sc(ws, r, 8, "YES" if moic >= moic_target else "NO", bold=True)
        if moic >= moic_target:
            meets_cell.fill = _GREEN_FILL
            meets_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        else:
            meets_cell.fill = _RED_FILL
            meets_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    # Property context
    r += 2
    _sc(ws, r, 1, "Property", bold=True)
    _sc(ws, r, 2, _safe(d, "property_name", ""))
    r += 1
    _sc(ws, r, 1, "Address", bold=True)
    _sc(ws, r, 2, _safe(d, "address", ""))
    r += 1
    _sc(ws, r, 1, "Units", bold=True)
    _sc(ws, r, 2, units)
    r += 1
    _sc(ws, r, 1, "T-12 NOI", bold=True)
    _sc(ws, r, 2, round(noi), "dollar")


def _remaining_balance(loan, rate, amort_years, years_paid):
    """Remaining loan balance after years_paid."""
    r = rate / 12
    n_total = amort_years * 12
    n_paid = years_paid * 12
    if r == 0:
        return loan * (1 - n_paid / n_total)
    monthly = loan * (r * (1 + r) ** n_total) / ((1 + r) ** n_total - 1)
    balance = loan * (1 + r) ** n_paid - monthly * ((1 + r) ** n_paid - 1) / r
    return max(0, balance)


# ─────────────────────────────────────────────────────────────────
# TAB 2 -- Sensitivity Pro Forma (monthly T-1..T-60)
# ─────────────────────────────────────────────────────────────────

def _build_sensitivity_proforma_tab(ws, d):
    ws.title = "Sensitivity Pro Forma"
    ws.column_dimensions["A"].width = 10
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 16

    noi = _safe(d, "noi", 0)
    monthly_noi = noi / 12
    debt_service = _safe(d, "annual_loan_payment", 0)
    monthly_ds = debt_service / 12
    rent_growth_monthly = (1.03) ** (1 / 12) - 1  # 3% annual -> monthly
    expense_growth_monthly = (1.03) ** (1 / 12) - 1

    egi = _safe(d, "effective_gross_income", 0)
    opex = _safe(d, "total_operating_expenses", 0)
    monthly_egi = egi / 12
    monthly_opex = opex / 12

    r = 1
    headers = ["Month", "Gross Income", "Operating Exp", "NOI", "Debt Service", "Cash Flow"]
    for c, hdr in enumerate(headers, 1):
        _sc(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, 6)

    # T-1 through T-60 (months 1-60)
    for month in range(1, 61):
        r += 1
        growth_factor = (1 + rent_growth_monthly) ** month
        exp_factor = (1 + expense_growth_monthly) ** month
        m_egi = monthly_egi * growth_factor
        m_opex = monthly_opex * exp_factor
        m_noi = m_egi - m_opex
        m_cf = m_noi - monthly_ds

        _sc(ws, r, 1, f"T-{month}", bold=(month % 12 == 0))
        _sc(ws, r, 2, round(m_egi), "dollar")
        _sc(ws, r, 3, round(m_opex), "dollar")
        _sc(ws, r, 4, round(m_noi), "dollar")
        _sc(ws, r, 5, round(monthly_ds), "dollar")
        _sc(ws, r, 6, round(m_cf), "dollar")


# ─────────────────────────────────────────────────────────────────
# TAB 3 -- Projections (year-by-year with MOIC)
# ─────────────────────────────────────────────────────────────────

def _build_projections_tab(ws, d):
    ws.title = "Projections"
    ws.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 18

    price = _safe(d, "price", 0)
    noi = _safe(d, "noi", 0)
    down_pct = _safe(d, "down_payment_pct", 0.20)
    dp = price * down_pct
    loan = price - dp
    interest_rate = _safe(d, "interest_rate", 0.075)
    debt_service = _safe(d, "annual_loan_payment", 0)
    exit_cap = _safe(d, "exit_cap_rate", 0.065)

    if debt_service == 0 and loan > 0:
        monthly_rate = interest_rate / 12
        n_payments = 30 * 12
        if monthly_rate > 0:
            monthly_pmt = loan * (monthly_rate * (1 + monthly_rate) ** n_payments) / (
                (1 + monthly_rate) ** n_payments - 1
            )
        else:
            monthly_pmt = loan / n_payments
        debt_service = monthly_pmt * 12

    r = 1
    headers = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for c, hdr in enumerate(headers, 1):
        _sc(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, 6)

    # Build rows
    noi_vals = []
    cf_vals = []
    cumulative_cf = []
    exit_vals = []
    moic_vals = []
    running_cf = 0

    for yr in range(1, 6):
        yr_noi = noi * (1.03 ** yr)
        yr_cf = yr_noi - debt_service
        running_cf += yr_cf
        yr_exit_noi = yr_noi
        yr_exit_val = yr_exit_noi / exit_cap if exit_cap > 0 else 0
        remaining = _remaining_balance(loan, interest_rate, 30, yr)
        net_exit = yr_exit_val * 0.96 - remaining
        yr_moic = (running_cf + net_exit) / dp if dp > 0 else 0

        noi_vals.append(yr_noi)
        cf_vals.append(yr_cf)
        cumulative_cf.append(running_cf)
        exit_vals.append(net_exit)
        moic_vals.append(yr_moic)

    row_data = [
        ("NOI", noi_vals, "dollar"),
        ("Debt Service", [debt_service] * 5, "dollar"),
        ("Annual Cash Flow", cf_vals, "dollar"),
        ("Cumulative Cash Flow", cumulative_cf, "dollar"),
        ("", [None] * 5, None),
        ("Exit Value (net of costs)", exit_vals, "dollar"),
        ("MOIC (cumul. CF + exit / equity)", moic_vals, "moic"),
    ]

    for label, vals, fmt in row_data:
        r += 1
        _sc(ws, r, 1, label, bold=True)
        for c, v in enumerate(vals, 2):
            if v is not None:
                _sc(ws, r, c, round(v, 2) if fmt == "moic" else round(v), fmt)

    r += 2
    _sc(ws, r, 1, "Total Equity Invested", bold=True)
    reno = _safe(d, "reno_cost_per_unit", 8000) * (_safe(d, "units", 1) or 1)
    _sc(ws, r, 2, round(dp + reno), "dollar")
    r += 1
    _sc(ws, r, 1, "Down Payment", bold=True)
    _sc(ws, r, 2, round(dp), "dollar")
    r += 1
    _sc(ws, r, 1, "Renovation Budget", bold=True)
    _sc(ws, r, 2, round(reno), "dollar")


# ─────────────────────────────────────────────────────────────────
# TAB 4 -- Comp table
# ─────────────────────────────────────────────────────────────────

def _build_comp_tab(ws, d):
    ws.title = "Comp"
    ws.column_dimensions["A"].width = 6
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 18

    r = 1
    headers = ["#", "Property Name", "Address", "Units", "Year Built",
               "Sale Price", "Price/Unit", "Cap Rate", "Sale Date"]
    for c, hdr in enumerate(headers, 1):
        _sc(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, len(headers))

    comps = _safe(d, "comps", [])
    for i in range(5):
        r += 1
        if i < len(comps):
            comp = comps[i]
            _sc(ws, r, 1, i + 1)
            _sc(ws, r, 2, comp.get("name", ""))
            _sc(ws, r, 3, comp.get("address", ""))
            _sc(ws, r, 4, comp.get("units", ""))
            _sc(ws, r, 5, comp.get("year_built", ""))
            _sc(ws, r, 6, comp.get("sale_price", ""), "dollar")
            _sc(ws, r, 7, comp.get("price_per_unit", ""), "dollar")
            _sc(ws, r, 8, comp.get("cap_rate", ""), "pct")
            _sc(ws, r, 9, comp.get("sale_date", ""))
        else:
            _sc(ws, r, 1, i + 1)
            for c in range(2, 10):
                _sc(ws, r, c, "")


# ─────────────────────────────────────────────────────────────────
# TAB 5 -- CapEx placeholder
# ─────────────────────────────────────────────────────────────────

def _build_capex_tab(ws, d):
    ws.title = "CapEx"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 30

    r = 1
    headers = ["Item", "Cost/Unit", "Total Cost", "Notes"]
    for c, hdr in enumerate(headers, 1):
        _sc(ws, r, c, hdr, bold=True)
    _style_header_row(ws, r, 4)

    units = _safe(d, "units", 1) or 1
    items = [
        ("Interior Renovations", 8000, "Flooring, paint, fixtures"),
        ("Exterior / Curb Appeal", 1500, "Paint, landscaping, signage"),
        ("HVAC", 1000, "As needed"),
        ("Plumbing", 500, "Unit + common area"),
        ("Roof Reserve", 800, "Pro-rated"),
    ]
    for name, cost, notes in items:
        r += 1
        _sc(ws, r, 1, name, bold=True)
        _sc(ws, r, 2, cost, "dollar")
        _sc(ws, r, 3, cost * units, "dollar")
        _sc(ws, r, 4, notes)

    r += 2
    _sc(ws, r, 1, "TOTAL", bold=True)
    total = sum(c for _, c, _ in items) * units
    _sc(ws, r, 3, total, "dollar")


# ─────────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────────

def generate(deal_data: dict, filepath: str) -> str:
    """
    Generate a Sensitivity Test xlsx workbook and save it to *filepath*.

    Parameters
    ----------
    deal_data : dict
        Property and financial data for the deal.
    filepath : str
        Destination path for the .xlsx file.

    Returns
    -------
    str  The filepath that was written.
    """
    wb = Workbook()

    _build_summary_tab(wb.active, deal_data)
    _build_sensitivity_proforma_tab(wb.create_sheet(), deal_data)
    _build_projections_tab(wb.create_sheet(), deal_data)
    _build_comp_tab(wb.create_sheet(), deal_data)
    _build_capex_tab(wb.create_sheet(), deal_data)

    wb.save(filepath)
    return filepath
